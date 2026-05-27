#!/usr/bin/env python3
"""Check blog-post data files against the master workbook compound list.

This script verifies that CSV, TSV, HTML, Markdown, Python, and workbook files
reference the compounds represented in the blog-post master workbook, plus
reference compounds needed for the displayed figures.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


OCNT_RE = re.compile(r"OCNT-\d+")
FIGURE_REFERENCE_IDS = {
    "OCNT-2317298",
}


@dataclass
class DelimitedFileReport:
    path: Path
    total_rows: int
    unexpected_rows: int
    unexpected_ids: set[str]


@dataclass
class AuditHit:
    path: Path
    ids: set[str]


def extract_canonical_ids(value: object) -> set[str]:
    if value is None:
        return set()
    return set(OCNT_RE.findall(str(value)))


def load_allowed_ids(master_workbook: Path) -> set[str]:
    workbook = load_workbook(master_workbook, read_only=True, data_only=True)
    allowed: set[str] = set()
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                allowed.update(extract_canonical_ids(value))
    allowed.update(FIGURE_REFERENCE_IDS)
    return allowed


def read_delimited(path: Path, delimiter: str) -> list[list[str]]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        return list(csv.reader(handle, delimiter=delimiter))


def check_delimited_file(path: Path, allowed_ids: set[str]) -> DelimitedFileReport:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    rows = read_delimited(path, delimiter)
    if not rows:
        return DelimitedFileReport(path, 0, 0, set())

    unexpected_ids: set[str] = set()
    unexpected_rows = 0

    for row in rows[1:]:
        row_ids = extract_canonical_ids("\t".join(row))
        if row_ids and not row_ids.issubset(allowed_ids):
            unexpected_rows += 1
            unexpected_ids.update(row_ids - allowed_ids)

    return DelimitedFileReport(
        path=path,
        total_rows=max(len(rows) - 1, 0),
        unexpected_rows=unexpected_rows,
        unexpected_ids=unexpected_ids,
    )


def iter_release_delimited_files(root: Path) -> Iterable[Path]:
    for base in (root / "data" / "raw", root / "data" / "processed"):
        if not base.exists():
            continue
        for suffix in ("*.csv", "*.tsv"):
            yield from sorted(base.rglob(suffix))


def iter_audit_text_files(root: Path) -> Iterable[Path]:
    suffixes = {".csv", ".tsv", ".json", ".html", ".md", ".py"}
    skip_dirs = {".git", "node_modules"}
    for path in sorted(root.rglob("*")):
        if not path.is_file():
            continue
        if any(part in skip_dirs for part in path.parts):
            continue
        if path.suffix.lower() in suffixes:
            yield path


def audit_text_file(path: Path, allowed_ids: set[str]) -> AuditHit | None:
    try:
        text = path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        text = path.read_text(encoding="utf-8", errors="ignore")
    ids = extract_canonical_ids(text) - allowed_ids
    if ids:
        return AuditHit(path, ids)
    return None


def audit_xlsx_file(path: Path, allowed_ids: set[str]) -> AuditHit | None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    ids: set[str] = set()
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                ids.update(extract_canonical_ids(value))
    ids -= allowed_ids
    if ids:
        return AuditHit(path, ids)
    return None


def audit_release_files(root: Path, allowed_ids: set[str]) -> list[AuditHit]:
    hits: list[AuditHit] = []
    for path in sorted(root.rglob("*")):
        if ".git" in path.parts:
            continue
        path_ids = extract_canonical_ids(path.relative_to(root).as_posix()) - allowed_ids
        if path_ids:
            hits.append(AuditHit(path, path_ids))
    for path in iter_audit_text_files(root):
        hit = audit_text_file(path, allowed_ids)
        if hit:
            hits.append(hit)
    for path in sorted((root / "data").rglob("*.xlsx")):
        hit = audit_xlsx_file(path, allowed_ids)
        if hit:
            hits.append(hit)
    return hits


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--repo-root", type=Path, default=Path.cwd())
    parser.add_argument("--json-report", type=Path)
    args = parser.parse_args()

    root = args.repo_root.resolve()
    master = root / "data" / "raw" / "master" / "PXR_CAD_Blog_Post_Master_Raw_Data.xlsx"
    if not master.exists():
        print(f"Master workbook not found: {master}", file=sys.stderr)
        return 2

    allowed_ids = load_allowed_ids(master)
    reports = [
        check_delimited_file(path, allowed_ids)
        for path in iter_release_delimited_files(root)
    ]
    unexpected_delimited = [report for report in reports if report.unexpected_rows]
    audit_hits = audit_release_files(root, allowed_ids)

    print(f"Expected blog-post compound IDs: {len(allowed_ids)}")
    print(f"Figure reference IDs: {', '.join(sorted(FIGURE_REFERENCE_IDS))}")
    print(f"Delimited files scanned: {len(reports)}")
    print(f"Delimited files with unexpected rows: {len(unexpected_delimited)}")
    for report in unexpected_delimited:
        rel = report.path.relative_to(root)
        print(
            f"- {rel}: unexpected rows {report.unexpected_rows}/{report.total_rows}; "
            f"unexpected IDs {len(report.unexpected_ids)}"
        )

    if audit_hits:
        print("Unexpected OCNT IDs are present:")
        for hit in audit_hits[:50]:
            rel = hit.path.relative_to(root)
            preview = ", ".join(sorted(hit.ids)[:12])
            more = "" if len(hit.ids) <= 12 else f" ... +{len(hit.ids) - 12}"
            print(f"- {rel}: {preview}{more}")
    else:
        print("Check passed: no unexpected OCNT IDs found.")

    if args.json_report:
        payload = {
            "allowed_id_count": len(allowed_ids),
            "unexpected_delimited_files": [
                {
                    "path": str(report.path.relative_to(root)),
                    "total_rows": report.total_rows,
                    "unexpected_rows": report.unexpected_rows,
                    "unexpected_id_count": len(report.unexpected_ids),
                    "unexpected_ids": sorted(report.unexpected_ids),
                }
                for report in unexpected_delimited
            ],
            "audit_hits": [
                {
                    "path": str(hit.path.relative_to(root)),
                    "id_count": len(hit.ids),
                    "ids": sorted(hit.ids),
                }
                for hit in audit_hits
            ],
        }
        args.json_report.parent.mkdir(parents=True, exist_ok=True)
        args.json_report.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")

    return 1 if unexpected_delimited or audit_hits else 0


if __name__ == "__main__":
    raise SystemExit(main())
