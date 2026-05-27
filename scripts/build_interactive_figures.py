#!/usr/bin/env python3
"""Build the interactive PXR blog-post figures.

The selected figure HTML files contain their Plotly runtime, styling, and
embedded hover assets. This script refreshes the embedded figure data from the
master workbook and DRC tables.
"""

from __future__ import annotations

import argparse
import base64
import copy
import csv
import html
import json
import math
import re
import statistics
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
FIGURES_DIR = REPO_ROOT / "figures" / "interactive"
FIGURE_DATA_DIR = REPO_ROOT / "data" / "processed" / "figure-payloads"
MASTER_WORKBOOK = REPO_ROOT / "data" / "raw" / "master" / "PXR_CAD_Blog_Post_Master_Raw_Data.xlsx"
MICROSCALE_DRC_DIR = (
    REPO_ROOT
    / "data"
    / "raw"
    / "drc"
    / "microscale-synthesis"
    / "OPA_PXR-048_56149_processed-data_96_semipures"
)

FIGURE_FILES = [
    "Figure 9A - Beeswarm.html",
    "Figure 9B - OCNT-2432456 Library.html",
    "Figure 9C - OCNT-0002416 Library.html",
    "Figure 9D - OCNT-2432456 Priority Rank.html",
    "Figure 9E - OCNT-0002416 Priority Rank.html",
]
BEESWARM_FILE = "Figure 9A - Beeswarm.html"

DATA_SCRIPT_RE = re.compile(
    r"(<script[^>]*\bid=[\"']dashboard-data[\"'][^>]*>)(.*?)(</script>)",
    re.IGNORECASE | re.DOTALL,
)

MICROSCALE_SHEET = "96-Compound \u00b5Scale Semi-Pure"
MICROSCALE_LABEL = "96-Compound Microscale Synthesis"
MICROSCALE_X = 2.0


def is_number(value: Any) -> bool:
    if value is None:
        return False
    try:
        number = float(value)
    except (TypeError, ValueError):
        return False
    return math.isfinite(number)


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in {"", "NA", "NaN", "nan", "no data"}:
            return None
        value = stripped.replace(",", "")
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    return number


def mean_numeric(values: Iterable[Any]) -> float | None:
    numbers = [to_float(value) for value in values]
    numbers = [value for value in numbers if value is not None]
    if not numbers:
        return None
    return statistics.mean(numbers)


def first_present(row: dict[str, Any], *names: str) -> Any:
    for name in names:
        if name in row:
            return row[name]
    return None


def p_ec50_from_um(ec50_um: float | None) -> float | None:
    if ec50_um is None or ec50_um <= 0:
        return None
    return 6 - math.log10(ec50_um)


def fmt_p(value: Any) -> str:
    number = to_float(value)
    return "no data" if number is None else f"{number:.2f}"


def fmt_yield(value: Any) -> str:
    number = to_float(value)
    return "no data" if number is None else f"{number:.1f}%"


def fmt_emax(value: Any) -> str:
    number = to_float(value)
    return "no data" if number is None else f"{number:,.0f}"


def yield_source_from_temperature(row: dict[str, Any], *names: str) -> str | None:
    temperature = to_float(first_present(row, *names))
    if temperature is None:
        return None
    return f"{temperature:g} °C EvapT"


def slug_from_filename(filename: str) -> str:
    return filename.removesuffix(".html")


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def write_text_if_changed(path: Path, value: str) -> bool:
    if path.exists() and path.read_text(encoding="utf-8") == value:
        return False
    path.write_text(value, encoding="utf-8")
    return True


def extract_payload_from_html(path: Path) -> dict[str, Any]:
    text = read_text(path)
    match = DATA_SCRIPT_RE.search(text)
    if not match:
        raise ValueError(f"Could not find dashboard-data script in {path}")
    payload_text = html.unescape(match.group(2).strip())
    return json.loads(payload_text)


def replace_payload_in_html(path: Path, payload: dict[str, Any]) -> bool:
    text = read_text(path)
    match = DATA_SCRIPT_RE.search(text)
    if not match:
        raise ValueError(f"Could not find dashboard-data script in {path}")
    payload_text = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    payload_text = payload_text.replace("</script", "<\\/script")
    updated = text[: match.start(2)] + payload_text + text[match.end(2) :]
    return write_text_if_changed(path, updated)


def write_payload_json(payload_dir: Path, filename: str, payload: dict[str, Any]) -> bool:
    payload_dir.mkdir(parents=True, exist_ok=True)
    target = payload_dir / f"{slug_from_filename(filename)}.json"
    value = json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True)
    value += "\n"
    return write_text_if_changed(target, value)


def parse_data_uri(uri: str) -> str:
    if not uri.startswith("data:"):
        raise ValueError("Expected a data URI")
    header, encoded = uri.split(",", 1)
    if ";base64" in header:
        return base64.b64decode(encoded).decode("utf-8")
    return encoded


def data_uri_svg(svg: str) -> str:
    encoded = base64.b64encode(svg.encode("utf-8")).decode("ascii")
    return f"data:image/svg+xml;base64,{encoded}"


def nested_svg_children(svg: str) -> str:
    match = re.search(r"<svg\b[^>]*>(.*)</svg>\s*$", svg, flags=re.DOTALL)
    return match.group(1) if match else svg


def text_svg(text: str, x: float, y: float, *, size: int = 12, weight: int = 500, anchor: str = "start", fill: str = "#4a5565") -> str:
    escaped = html.escape(text)
    return (
        f'<text x="{x}" y="{y}" font-size="{size}" font-weight="{weight}" '
        f'text-anchor="{anchor}" fill="{fill}" '
        'font-family="Inter, Arial, sans-serif">'
        f"{escaped}</text>"
    )


def make_summary_svg(point: dict[str, Any]) -> str:
    is_missing = point.get("adjStatus") in {"Undetected on CAD", "Below CAD LLOQ"}
    height = 110 if is_missing else 184
    width = 620
    role_label = point.get("roleLabel", "Compound")
    role_fill = "#2f6ca3" if point.get("role") == "hit" else "#3b8b68"

    parts: list[str] = [
        (
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" '
            f'viewBox="0 0 {width} {height}">'
        ),
        '<rect x="0" y="0" width="620" height="{height}" rx="8" fill="#ffffff" stroke="#d8e1ec" stroke-width="1.5"/>'.format(
            height=height
        ),
        text_svg(point["sampleId"], 24, 30, size=18, weight=800, fill="#202938"),
        f'<rect x="398" y="12" width="198" height="34" rx="17" fill="{role_fill}"/>',
        text_svg(role_label, 497, 35, size=13, weight=700, anchor="middle", fill="#ffffff"),
    ]

    if is_missing:
        status = point.get("adjStatus", "No adjusted pEC50")
        detail = "No numeric yield-adjusted potency or CAD-derived assay summary was assigned."
        parts.extend(
            [
                text_svg(status, 24, 70, size=16, weight=800, fill="#27364a"),
                text_svg(detail, 24, 95, size=12, weight=600, fill="#657287"),
            ]
        )
        parts.append("</svg>")
        return "".join(parts)

    raw = to_float(point.get("rawPEc50"))
    adj = to_float(point.get("adjPEc50"))
    raw_x = 130 + max(0, min(1, ((raw or 1) - 1) / 6)) * 430
    adj_x = 130 + max(0, min(1, ((adj or 1) - 1) / 6)) * 430
    delta = (adj - raw) if raw is not None and adj is not None else None

    parts.extend(
        [
            text_svg(f"QC: {point.get('qcFlag', 'PASS')}", 24, 66, size=13, weight=700),
            text_svg("Activity: YES", 205, 66, size=13, weight=700),
            text_svg("Potency", 24, 105, size=14, weight=500),
            '<line x1="130" y1="96" x2="560" y2="96" stroke="#c9c9c9" stroke-width="2"/>',
        ]
    )
    for tick in range(1, 8):
        x = 130 + ((tick - 1) / 6) * 430
        parts.append(f'<line x1="{x}" y1="86" x2="{x}" y2="106" stroke="#b4b4b4" stroke-width="1.5"/>')
        parts.append(text_svg(str(tick), x, 80, size=12, weight=500, anchor="middle"))
    parts.extend(
        [
            f'<circle cx="{raw_x}" cy="96" r="8" fill="#e15454" stroke="#ffffff" stroke-width="2"/>',
            f'<rect x="{adj_x - 8}" y="88" width="16" height="16" fill="#4d7fb1" stroke="#ffffff" stroke-width="2"/>',
            text_svg(f"Uncorr. {fmt_p(raw)}", raw_x, 126, size=12, weight=600, anchor="middle", fill="#e15454"),
            text_svg(f"Adj. {fmt_p(adj)}", adj_x, 150, size=12, weight=600, anchor="middle", fill="#4d7fb1"),
            text_svg(f"Δ pEC50 {delta:+.2f}" if delta is not None else "Δ pEC50 no data", 560, 136, size=13, weight=600, anchor="end"),
            text_svg("Yield", 58, 176, size=14, weight=500, anchor="middle"),
            '<rect x="152" y="166" width="408" height="16" fill="#f0f2f4" stroke="#d9dde2"/>',
        ]
    )

    yld = to_float(point.get("cadYield"))
    if yld is not None:
        yield_width = max(0, min(408, (min(yld, 100) / 100) * 408))
        parts.append(f'<rect x="152" y="166" width="{yield_width}" height="16" fill="#77bdb8"/>')
    parts.append(text_svg(fmt_yield(yld), 560, 178, size=12, weight=600, anchor="end"))
    parts.append("</svg>")
    return "".join(parts)


def scale_x(value: float, x_min: float, x_max: float, left: float = 92, right: float = 574) -> float:
    value = max(x_min, min(x_max, value))
    return left + ((value - x_min) / (x_max - x_min)) * (right - left)


def scale_y(value: float, y_min: float, y_max: float, top: float = 36, bottom: float = 316) -> float:
    value = max(y_min, min(y_max, value))
    return bottom - ((value - y_min) / (y_max - y_min)) * (bottom - top)


def make_drc_svg(point: dict[str, Any], concentration_rows: list[dict[str, str]], drc_rows: list[dict[str, str]]) -> str | None:
    if point.get("adjStatus") in {"Undetected on CAD", "Below CAD LLOQ"}:
        return None
    rows = concentration_rows
    if not rows:
        return None

    concentrations = [to_float(row.get("concentration_uM")) for row in rows]
    signals = [to_float(row.get("norm")) for row in rows]
    pairs = [(c, s) for c, s in zip(concentrations, signals) if c is not None and c > 0 and s is not None]
    if not pairs:
        return None

    max_signal = max(1.2, max(signal for _, signal in pairs) + 0.15)
    if point.get("sampleId") == "OCNT-2317298":
        max_signal = 1.5

    x_min, x_max = math.log10(0.005), math.log10(50)
    y_min, y_max = -0.25, max_signal
    width, height = 660, 390

    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        '<rect x="0" y="0" width="660" height="390" rx="8" fill="#ffffff" stroke="#d8e1ec" stroke-width="1.5"/>',
        '<rect x="92" y="36" width="482" height="280" fill="#ffffff" stroke="#d9d9d9" stroke-width="1"/>',
    ]

    for y_tick in [0, 0.5, 1.0, 1.5]:
        if y_tick > y_max:
            continue
        y = scale_y(y_tick, y_min, y_max)
        parts.append(f'<line x1="92" x2="574" y1="{y}" y2="{y}" stroke="#e4e8ee" stroke-dasharray="5 5"/>')
        parts.append(text_svg(f"{y_tick:.1f}", 80, y + 4, size=11, weight=500, anchor="end"))

    x_ticks = [0.005, 0.05, 0.5, 5, 50]
    for x_tick in x_ticks:
        x = scale_x(math.log10(x_tick), x_min, x_max)
        parts.append(f'<line x1="{x}" x2="{x}" y1="316" y2="326" stroke="#818181"/>')
        parts.append(text_svg(f"{x_tick:g}", x, 344, size=11, weight=500, anchor="middle"))

    points = [
        (
            scale_x(math.log10(conc), x_min, x_max),
            scale_y(signal, y_min, y_max),
            conc,
            signal,
        )
        for conc, signal in pairs
    ]
    points.sort(key=lambda item: item[2])

    fit_row = drc_rows[0] if drc_rows else {}
    raw_ec50 = to_float(fit_row.get("ec50")) or to_float(point.get("rawEC50"))
    if raw_ec50 is not None and raw_ec50 > 0:
        x = scale_x(math.log10(raw_ec50), x_min, x_max)
        parts.append(f'<line x1="{x}" x2="{x}" y1="42" y2="316" stroke="#e95b5b" stroke-width="2" stroke-dasharray="7 5"/>')
        parts.append(text_svg("Uncorrected EC50", min(x + 6, 570), 62, size=11, weight=700, anchor="end", fill="#e84e4e"))

    if len(points) > 1:
        path = " ".join(f"{'M' if idx == 0 else 'L'} {x:.2f} {y:.2f}" for idx, (x, y, _, _) in enumerate(points))
        parts.append(f'<path d="{path}" fill="none" stroke="#2f80c2" stroke-width="3"/>')
    for x, y, _, _ in points:
        parts.append(f'<circle cx="{x}" cy="{y}" r="5.5" fill="#f28e2b" stroke="#ffffff" stroke-width="1"/>')

    structure_uri = point.get("structureImage")
    if structure_uri:
        structure_svg = parse_data_uri(structure_uri)
        children = nested_svg_children(structure_svg)
        x_pos, y_pos, box_w, box_h = 64, 76, 96, 50
        if point.get("sampleId") == "OCNT-2308748":
            x_pos = 88
        if point.get("sampleId") == "OCNT-2317298":
            y_pos = 118
        parts.append(f'<svg x="{x_pos}" y="{y_pos}" width="{box_w}" height="{box_h}" viewBox="0 0 420 260">{children}</svg>')

    parts.extend(
        [
            '<text x="22" y="190" transform="rotate(-90 22 190)" font-size="11" font-weight="600" text-anchor="middle" fill="#4a5565" font-family="Inter, Arial, sans-serif">Luciferase signal (normalized to positive control)</text>',
            text_svg("Unadjusted Concentration (µM)", 333, 376, size=12, weight=600, anchor="middle"),
            "</svg>",
        ]
    )
    return "".join(parts)


def load_workbook_rows(sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(MASTER_WORKBOOK, data_only=False, read_only=True)
    try:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    output: list[dict[str, Any]] = []
    for row in rows[1:]:
        item = {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}
        if any(value is not None and value != "" for value in item.values()):
            output.append(item)
    return output


def htchem_sample_id(row: dict[str, Any]) -> str:
    compound = str(row.get("Compound ID") or "").strip()
    batch = str(row.get("Crude Batch ID") or "").strip()
    if not compound:
        raise ValueError("HTChem row is missing Compound ID")
    if compound.endswith("-001") or not batch:
        return compound
    return f"{compound}-{batch}"


def htchem_status_from_master(row: dict[str, Any]) -> tuple[float | None, str | None, str | None]:
    peak_area_value = row.get("Crude Peak Area (pA*min)")
    theoretical_mass = to_float(row.get("Crude Theoretical Mass-on-Column (ng)"))
    status_text = str(peak_area_value).strip() if peak_area_value is not None else ""
    if status_text == "Undetected":
        return None, "Undetected on CAD", None
    if status_text == "Below LLOQ":
        return None, "Below CAD LLOQ", None
    peak_area = to_float(peak_area_value)
    if peak_area is None or theoretical_mass is None or theoretical_mass <= 0:
        return None, "No adjusted pEC50", None
    return (
        12.5 * peak_area / theoretical_mass * 100.0,
        None,
        yield_source_from_temperature(row, "Crude EvapT (°C)"),
    )


def build_master_truth() -> dict[str, dict[str, Any]]:
    truth: dict[str, dict[str, Any]] = {}

    for row in load_workbook_rows("HTChem Libraries"):
        sample_id = htchem_sample_id(row)
        raw_ec50 = to_float(row.get("Crude EC50s (µM)"))
        raw_pec50 = p_ec50_from_um(raw_ec50)
        cad_yield, adj_status, yield_source = htchem_status_from_master(row)
        adj_ec50 = raw_ec50 * cad_yield / 100.0 if raw_ec50 is not None and cad_yield is not None else None
        truth[sample_id] = {
            "rawPEc50": raw_pec50,
            "adjPEc50": p_ec50_from_um(adj_ec50),
            "cadYield": cad_yield,
            "adjStatus": adj_status,
            "yieldSource": yield_source,
        }

    for row in load_workbook_rows(MICROSCALE_SHEET):
        sample_id = sample_id_from_microscale_row(row)
        raw_ec50 = to_float(row.get("Semi-Pure EC50s (µM)"))
        raw_pec50 = p_ec50_from_um(raw_ec50)
        if raw_pec50 is None:
            continue
        cad_yield, adj_status, yield_source = microscale_status_from_master(row)
        adj_ec50 = raw_ec50 * cad_yield / 100.0 if cad_yield is not None else None
        truth[sample_id] = {
            "rawPEc50": raw_pec50,
            "adjPEc50": p_ec50_from_um(adj_ec50),
            "cadYield": cad_yield,
            "adjStatus": adj_status,
            "yieldSource": yield_source,
        }

    return truth


def close_enough(actual: Any, expected: Any, tolerance: float = 0.02) -> bool:
    actual_number = to_float(actual)
    expected_number = to_float(expected)
    if actual_number is None and expected_number is None:
        return True
    if actual_number is None or expected_number is None:
        return False
    return abs(actual_number - expected_number) <= tolerance


def validate_payload_values_against_master(payloads: dict[str, dict[str, Any]]) -> None:
    truth = build_master_truth()
    errors: list[str] = []
    fields = ("rawPEc50", "adjPEc50", "cadYield")

    for filename, payload in payloads.items():
        for point in payload.get("points", []):
            sample_id = point.get("sampleId")
            if not sample_id or sample_id not in truth:
                continue
            expected = truth[sample_id]
            for field in fields:
                if not close_enough(point.get(field), expected.get(field)):
                    errors.append(
                        f"{filename}: {sample_id} {field} is {point.get(field)!r}, "
                        f"expected {expected.get(field)!r}"
                    )
            expected_status = expected.get("adjStatus")
            if expected_status and point.get("adjStatus") != expected_status:
                errors.append(
                    f"{filename}: {sample_id} adjStatus is {point.get('adjStatus')!r}, "
                    f"expected {expected_status!r}"
                )
            expected_source = expected.get("yieldSource")
            actual_source = point.get("yieldSource")
            if expected_source is None:
                if actual_source not in (None, "", "no data"):
                    errors.append(
                        f"{filename}: {sample_id} yieldSource is {actual_source!r}, expected no value"
                    )
            elif actual_source != expected_source:
                errors.append(
                    f"{filename}: {sample_id} yieldSource is {actual_source!r}, "
                    f"expected {expected_source!r}"
                )

    if errors:
        preview = "\n".join(errors[:20])
        more = "" if len(errors) <= 20 else f"\n... and {len(errors) - 20} more"
        raise ValueError(f"Figure payload values do not match the master workbook:\n{preview}{more}")


def apply_master_truth_to_payloads(payloads: dict[str, dict[str, Any]]) -> None:
    truth = build_master_truth()
    for payload in payloads.values():
        for point in payload.get("points", []):
            sample_id = point.get("sampleId")
            if not sample_id or sample_id not in truth:
                continue
            expected = truth[sample_id]
            point["rawPEc50"] = expected.get("rawPEc50")
            point["rawLabel"] = fmt_p(expected.get("rawPEc50"))
            point["adjPEc50"] = expected.get("adjPEc50")
            point["adjPlotValue"] = expected.get("adjPEc50") if expected.get("adjPEc50") is not None else point.get("adjPlotValue")
            point["adjLabel"] = fmt_p(expected.get("adjPEc50"))
            point["cadYield"] = expected.get("cadYield")
            point["cadYieldCapped"] = min(expected["cadYield"], 100.0) if expected.get("cadYield") is not None else None
            point["yieldLabel"] = fmt_yield(expected.get("cadYield"))
            point["adjStatus"] = expected.get("adjStatus")
            point["yieldSource"] = expected.get("yieldSource")
            if expected.get("adjStatus") in {"Undetected on CAD", "Below CAD LLOQ", "No adjusted pEC50"}:
                point["adjLabel"] = "no data"
                point["cadYield"] = None
                point["cadYieldCapped"] = None
                point["yieldLabel"] = expected["adjStatus"]
                point["yieldSource"] = None


def load_stats_points(path: Path) -> dict[str, list[dict[str, str]]]:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            if row.get("compound_class") and row.get("compound_class") != "Library":
                continue
            sample_id = row.get("molecule_name", "").strip()
            if sample_id:
                grouped[sample_id].append(row)
    return grouped


def load_drc_rows(path: Path) -> dict[str, list[dict[str, str]]]:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            if row.get("compound_class") and row.get("compound_class") != "Library":
                continue
            sample_id = row.get("molecule_name", "").strip()
            if sample_id:
                grouped[sample_id].append(row)
    return grouped


def sample_id_from_microscale_row(row: dict[str, Any]) -> str:
    compound = str(row.get("Compound ID") or "").strip()
    batch = str(row.get("Semi-Pure Batch ID") or "").strip()
    if not compound:
        raise ValueError("Microscale row is missing Compound ID")
    if compound.endswith("-001") or not batch:
        return compound
    return f"{compound}-{batch}"


def microscale_status_from_master(row: dict[str, Any]) -> tuple[float | None, str | None, str | None]:
    peak_area_value = first_present(row, "Semi-Pure Product Peak Area", "Semi-Pure Peak Area (pA*min)")
    theoretical_mass = to_float(
        first_present(row, "Theoretical Product Mass (ng)", "Semi-Pure Theoretical Mass-on-Column (ng)")
    )
    status_text = str(peak_area_value).strip() if peak_area_value is not None else ""
    if status_text == "Undetected":
        return None, "Undetected on CAD", None
    if status_text == "Below LLOQ":
        return None, "Below CAD LLOQ", None
    peak_area = to_float(peak_area_value)
    if peak_area is None or theoretical_mass is None or theoretical_mass <= 0:
        return None, "No adjusted pEC50", "No adjusted pEC50"
    cad_yield = 12.5 * peak_area / theoretical_mass * 100.0
    return cad_yield, None, yield_source_from_temperature(row, "Semi-pure EvapT (°C)")


def fan_no_adjusted_rows(points: list[dict[str, Any]]) -> None:
    missing = [point for point in points if point.get("adjStatus") in {"Undetected on CAD", "Below CAD LLOQ", "No adjusted pEC50"}]
    missing.sort(key=lambda item: item.get("sampleId", ""))
    rows = [-0.12, -0.06, 0.0, 0.06, 0.12]
    for index, point in enumerate(missing):
        point["adjPlotValue"] = rows[index % len(rows)]


def assign_beeswarm_x(points: list[dict[str, Any]], base_x: float, *, min_gap: float = 0.075) -> None:
    sorted_points = sorted(points, key=lambda item: (to_float(item.get("adjPlotValue")) or 0, item.get("sampleId", "")))
    placed: list[dict[str, Any]] = []
    offsets = [0.0]
    for step in range(1, 24):
        offsets.extend([step * min_gap, -step * min_gap])
    for point in sorted_points:
        y = to_float(point.get("adjPlotValue")) or 0.0
        offset_choice = offsets[-1]
        for offset in offsets:
            ok = True
            for other in placed:
                other_y = to_float(other.get("adjPlotValue")) or 0.0
                dy = abs(y - other_y)
                if dy > 0.18:
                    continue
                dx = abs((base_x + offset) - float(other.get("libraryX", base_x)))
                required = max(0.03, min_gap * (1.0 - dy / 0.18))
                if dx < required:
                    ok = False
                    break
            if ok:
                offset_choice = offset
                break
        point["libraryX"] = round(base_x + offset_choice, 4)
        placed.append(point)


def build_microscale_points(existing_beeswarm: dict[str, Any]) -> list[dict[str, Any]]:
    rows = load_workbook_rows(MICROSCALE_SHEET)
    stats_points = load_stats_points(MICROSCALE_DRC_DIR / "stats.tsv")
    drc_rows = load_drc_rows(MICROSCALE_DRC_DIR / "drc.csv")
    cached = {
        point["sampleId"]: point
        for point in existing_beeswarm.get("points", [])
        if point.get("libraryNumber") == 3 or point.get("libraryLabel") == MICROSCALE_LABEL
    }

    points: list[dict[str, Any]] = []
    for row in rows:
        sample_id = sample_id_from_microscale_row(row)
        raw_ec50 = to_float(row.get("Semi-Pure EC50s (µM)"))
        raw_pec50 = p_ec50_from_um(raw_ec50)
        if raw_pec50 is None:
            continue
        if sample_id not in cached:
            raise ValueError(
                f"{sample_id} has raw pEC50 data but no embedded structure asset in the beeswarm HTML"
            )

        cad_yield, adj_status, yield_source = microscale_status_from_master(row)
        adj_ec50 = raw_ec50 * cad_yield / 100.0 if cad_yield is not None else None
        adj_pec50 = p_ec50_from_um(adj_ec50)
        adj_plot_value = adj_pec50 if adj_pec50 is not None else 0.0
        row_stats = stats_points.get(sample_id, [])
        row_fits = drc_rows.get(sample_id, [])
        cached_point = copy.deepcopy(cached[sample_id])
        concentration_norm = mean_numeric([item.get("norm") for item in row_stats])
        raw_emax = to_float(first_present(row, "Raw Emax", "Semi-Pure Emax raw"))
        normalized_emax = to_float(first_present(row, "Normalized Emax", "Semi-Pure Emax normalized"))
        qc_flags = [fit.get("qc_flag") for fit in row_fits if fit.get("qc_flag")]
        qc_flag = "PASS" if "PASS" in qc_flags else (qc_flags[0] if qc_flags else "PASS")

        point = cached_point
        point.update(
            {
                "sampleId": sample_id,
                "smiles": row.get("SMILES"),
                "role": "microscale",
                "roleLabel": "96-Compound Microscale",
                "activityStatus": "YES",
                "qcFlag": qc_flag,
                "rawEC50": raw_ec50,
                "rawPEc50": raw_pec50,
                "rawLabel": fmt_p(raw_pec50),
                "adjEC50": adj_ec50,
                "adjPEc50": adj_pec50,
                "adjPlotValue": adj_plot_value,
                "adjLabel": fmt_p(adj_pec50),
                "adjStatus": adj_status,
                "cadYield": cad_yield,
                "cadYieldCapped": min(cad_yield, 100.0) if cad_yield is not None else None,
                "yieldLabel": fmt_yield(cad_yield),
                "yieldSource": yield_source,
                "rawEmax": raw_emax,
                "normalizedEmax": normalized_emax,
                "emax": raw_emax,
                "emaxLabel": fmt_emax(raw_emax),
                "libraryNumber": 3,
                "libraryLabel": MICROSCALE_LABEL,
                "libraryX": MICROSCALE_X,
                "isReference": False,
                "isMissingAdjusted": adj_pec50 is None,
                "drcVisible": bool(row_stats) and adj_status is None,
                "concentrationCount": len(row_stats),
                "meanNormalizedSignal": concentration_norm,
            }
        )
        if adj_status in {"Undetected on CAD", "Below CAD LLOQ"}:
            point["cadYield"] = None
            point["cadYieldCapped"] = None
            point["yieldLabel"] = adj_status
            point["yieldSource"] = None
            point["drcVisible"] = False
            point["drcImage"] = None

        point["summaryImage"] = data_uri_svg(make_summary_svg(point))
        drc_svg = make_drc_svg(point, row_stats, row_fits)
        point["drcImage"] = data_uri_svg(drc_svg) if drc_svg else None
        points.append(point)

    fan_no_adjusted_rows(points)
    assign_beeswarm_x(points, MICROSCALE_X)
    return points


def rebuild_beeswarm_payload(existing: dict[str, Any]) -> dict[str, Any]:
    payload = copy.deepcopy(existing)
    base_points = [
        point
        for point in payload.get("points", [])
        if point.get("libraryNumber") != 3 and point.get("libraryLabel") != MICROSCALE_LABEL
    ]
    for point in base_points:
        cad_yield = to_float(point.get("cadYield"))
        if cad_yield is not None and to_float(point.get("cadYieldCapped")) is None:
            point["cadYieldCapped"] = min(cad_yield, 100.0)
    microscale_points = build_microscale_points(existing)
    payload["points"] = base_points + microscale_points

    chart = payload.setdefault("chart", {})
    chart.update(
        {
            "xLabel": "",
            "xRange": [-0.55, 2.55],
            "xTickValues": [0, 1, 2],
            "xTickText": [
                "OCNT-2432456<br>Core Library",
                "OCNT-0002416<br>Core Library",
                "96-Compound<br>Microscale Synthesis",
            ],
            "xTickMobileText": ["1", "2", "3"],
            "libraryLabel": "PXR HTChem Libraries and 96-Compound Microscale Synthesis",
            "title": "HTChem Beeswarm Plot",
            "subtitle": "",
        }
    )
    numeric_values = [
        to_float(point.get("adjPlotValue"))
        for point in payload["points"]
        if to_float(point.get("adjPlotValue")) is not None
    ]
    reference_values = [
        to_float(point.get("y"))
        for point in payload.get("referencePoints", [])
        if to_float(point.get("y")) is not None
    ]
    all_values = [value for value in numeric_values + reference_values if value is not None]
    if all_values:
        chart["yRange"] = [min(-0.45, min(all_values) - 0.2), max(7.25, max(all_values) + 0.25)]

    counts = defaultdict(int)
    for point in payload["points"]:
        counts[int(point.get("libraryNumber", 0))] += 1
    payload["summaryStats"] = {
        "totalCompounds": len(payload["points"]),
        "library1Compounds": counts[1],
        "library2Compounds": counts[2],
        "microscaleCompounds": counts[3],
        "numericAdjusted": sum(1 for point in payload["points"] if to_float(point.get("adjPEc50")) is not None),
        "notDetectedOnCad": sum(
            1
            for point in payload["points"]
            if point.get("adjStatus") in {"Undetected on CAD", "Below CAD LLOQ", "No adjusted pEC50"}
        ),
    }
    return payload


def validate_beeswarm_payload(payload: dict[str, Any]) -> None:
    points = payload.get("points", [])
    counts = defaultdict(int)
    sample_ids = set()
    for point in points:
        sample_id = point.get("sampleId")
        if not sample_id:
            raise ValueError("Beeswarm point missing sampleId")
        sample_ids.add(sample_id)
        counts[int(point.get("libraryNumber", 0))] += 1
        if point.get("cadYield") is not None and to_float(point.get("cadYieldCapped")) is None:
            raise ValueError(f"{sample_id} has CAD yield but no capped color value")

    expected = {1: 207, 2: 224, 3: 95}
    for library_number, count in expected.items():
        if counts[library_number] != count:
            raise ValueError(f"Expected library {library_number} count {count}, found {counts[library_number]}")
    if len(points) != sum(expected.values()):
        raise ValueError(f"Expected {sum(expected.values())} beeswarm points, found {len(points)}")
    if "OCNT-2469048-TL-001" in sample_ids:
        raise ValueError("OCNT-2469048-TL-001 should not appear because it has no usable DRC fit")
    undetected = next((point for point in points if point.get("sampleId") == "OCNT-2469069-TL-001"), None)
    if not undetected:
        raise ValueError("OCNT-2469069-TL-001 should appear as an Undetected on CAD microscale compound")
    if undetected.get("adjStatus") != "Undetected on CAD" or undetected.get("drcImage"):
        raise ValueError("OCNT-2469069-TL-001 should have the compact Undetected on CAD hover state")


def build_all(figures_dir: Path, payload_dir: Path | None) -> None:
    payloads: dict[str, dict[str, Any]] = {}
    for filename in FIGURE_FILES:
        path = figures_dir / filename
        payloads[filename] = extract_payload_from_html(path)

    payloads[BEESWARM_FILE] = rebuild_beeswarm_payload(payloads[BEESWARM_FILE])
    apply_master_truth_to_payloads(payloads)
    validate_payload_values_against_master(payloads)
    validate_beeswarm_payload(payloads[BEESWARM_FILE])

    changed_html = []
    changed_payloads = []
    for filename, payload in payloads.items():
        if replace_payload_in_html(figures_dir / filename, payload):
            changed_html.append(filename)
        if payload_dir is not None and write_payload_json(payload_dir, filename, payload):
            changed_payloads.append(f"{slug_from_filename(filename)}.json")

    print("Interactive figure build complete.")
    print(f"HTML files checked: {len(FIGURE_FILES)}")
    print(f"HTML files updated: {len(changed_html)}")
    if changed_html:
        for filename in changed_html:
            print(f"  - {filename}")
    if payload_dir is not None:
        print(f"Figure-data JSON files written under: {payload_dir}")
        print(f"Figure-data JSON files updated: {len(changed_payloads)}")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--figures-dir", type=Path, default=FIGURES_DIR, help="Directory containing interactive HTML files")
    parser.add_argument(
        "--figure-data-dir",
        type=Path,
        default=None,
        help="Optional directory for generated figure-data JSON.",
    )
    parser.add_argument("--payload-dir", dest="figure_data_dir", type=Path, help=argparse.SUPPRESS)
    parser.add_argument("--no-payload-json", action="store_true", help=argparse.SUPPRESS)
    args = parser.parse_args()

    figure_data_dir = None if args.no_payload_json else args.figure_data_dir
    build_all(args.figures_dir, figure_data_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
